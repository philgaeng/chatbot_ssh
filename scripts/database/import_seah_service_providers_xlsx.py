#!/usr/bin/env python3
"""Import SEAH service providers from the KL Road Excel workbook into Postgres.

Usage (from repo root):
    python scripts/database/import_seah_service_providers_xlsx.py \\
        --xlsx backend/dev-resources/SEAH\\ Service\\ Providers_NEP.xlsx \\
        [--csv scripts/database/seeds/seah_service_providers_kl_road.csv] \\
        [--dry-run]
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]
sys.path.insert(0, str(PROJECT_ROOT))
os.environ.setdefault("PYTHONPATH", str(PROJECT_ROOT))

from backend.shared_functions.seah_service_providers import (  # noqa: E402
    _clean_text,
    _slug,
    resolve_location_codes,
)

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required: pip install openpyxl") from exc

CSV_FIELDS = [
    "seah_service_provider_id",
    "country_code",
    "province_code",
    "district_code",
    "municipality_code",
    "province",
    "district",
    "municipality",
    "ward",
    "seah_center_name",
    "address",
    "phone",
    "opening_days",
    "opening_hours",
    "remarks",
    "is_active",
    "sort_order",
]


def _cell(value: Any) -> str:
    if value is None:
        return ""
    return _clean_text(unicodedata.normalize("NFKC", str(value)))


def _make_provider_id(
    district_code: Optional[str],
    municipality_code: Optional[str],
    ward: str,
    center_name: str,
    seq: int,
) -> str:
    scope = municipality_code or district_code or "np"
    parts = [_slug(scope), _slug(center_name, 32)]
    if ward:
        parts.append(f"w{ward}")
    parts.append(str(seq))
    return "-".join(parts)


def parse_workbook(xlsx_path: Path) -> List[Dict[str, str]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    rows: List[Dict[str, str]] = []
    global_seq = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_rows = list(ws.iter_rows(values_only=True))
        if len(sheet_rows) < 2:
            continue

        district_fallback = _clean_text(sheet_name)
        for raw in sheet_rows[2:]:
            if not raw or len(raw) < 7 or raw[1] in (None, "", "s. no. "):
                continue

            def col(idx: int) -> str:
                return _cell(raw[idx]) if idx < len(raw) else ""

            province = col(2) or "Koshi"
            district = col(3) or district_fallback
            municipality = col(4)
            ward = col(5)
            center_name = col(6)
            if not center_name:
                continue

            address = col(7)
            phone = col(8)
            opening_days = col(9)
            opening_hours = col(10)
            remarks = col(11)

            codes = resolve_location_codes(province, district, municipality or None)
            global_seq += 1
            provider_id = _make_provider_id(
                codes.get("district_code"),
                codes.get("municipality_code"),
                ward,
                center_name,
                global_seq,
            )

            rows.append(
                {
                    "seah_service_provider_id": provider_id,
                    "country_code": codes.get("country_code") or "NP",
                    "province_code": codes.get("province_code") or "",
                    "district_code": codes.get("district_code") or "",
                    "municipality_code": codes.get("municipality_code") or "",
                    "province": province,
                    "district": district,
                    "municipality": municipality,
                    "ward": ward,
                    "seah_center_name": center_name,
                    "address": address,
                    "phone": phone,
                    "opening_days": opening_days,
                    "opening_hours": opening_hours,
                    "remarks": remarks,
                    "is_active": "true",
                    "sort_order": str(global_seq),
                }
            )
    return rows


def write_csv(rows: List[Dict[str, str]], csv_path: Path) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def read_csv(csv_path: Path) -> List[Dict[str, str]]:
    if not csv_path.is_file():
        raise FileNotFoundError(csv_path)
    with csv_path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def upsert_rows(rows: List[Dict[str, str]]) -> int:
    from backend.services.database_services.postgres_services import db_manager

    sql = """
        INSERT INTO seah_service_providers (
            seah_service_provider_id, country_code, province_code, district_code,
            municipality_code, province, district, municipality, ward,
            seah_center_name, address, phone, opening_days, opening_hours,
            remarks, is_active, sort_order
        ) VALUES (
            %(seah_service_provider_id)s, %(country_code)s, %(province_code)s, %(district_code)s,
            %(municipality_code)s, %(province)s, %(district)s, %(municipality)s, %(ward)s,
            %(seah_center_name)s, %(address)s, %(phone)s, %(opening_days)s, %(opening_hours)s,
            %(remarks)s, %(is_active)s, %(sort_order)s
        )
        ON CONFLICT (seah_service_provider_id) DO UPDATE SET
            country_code = EXCLUDED.country_code,
            province_code = EXCLUDED.province_code,
            district_code = EXCLUDED.district_code,
            municipality_code = EXCLUDED.municipality_code,
            province = EXCLUDED.province,
            district = EXCLUDED.district,
            municipality = EXCLUDED.municipality,
            ward = EXCLUDED.ward,
            seah_center_name = EXCLUDED.seah_center_name,
            address = EXCLUDED.address,
            phone = EXCLUDED.phone,
            opening_days = EXCLUDED.opening_days,
            opening_hours = EXCLUDED.opening_hours,
            remarks = EXCLUDED.remarks,
            is_active = EXCLUDED.is_active,
            sort_order = EXCLUDED.sort_order,
            updated_at = CURRENT_TIMESTAMP
    """
    count = 0
    for row in rows:
        payload = dict(row)
        payload["is_active"] = str(row.get("is_active", "true")).lower() in {"1", "true", "yes", "y"}
        payload["sort_order"] = int(row.get("sort_order") or 0)
        for code_field in ("province_code", "district_code", "municipality_code"):
            if not payload.get(code_field):
                payload[code_field] = None
        db_manager.execute_update(sql, payload)
        count += 1
    return count


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xlsx",
        default=str(PROJECT_ROOT / "backend/dev-resources/SEAH Service Providers_NEP.xlsx"),
    )
    parser.add_argument(
        "--csv",
        default=str(SCRIPT_DIR / "seeds/seah_service_providers_kl_road.csv"),
    )
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--from-csv",
        action="store_true",
        help="Upsert from --csv only (skip Excel parse). Use for demo/prod seeding.",
    )
    args = parser.parse_args()

    csv_path = Path(args.csv)

    if args.from_csv:
        rows = read_csv(csv_path)
        if not rows:
            raise SystemExit(f"No provider rows in CSV: {csv_path}")
        print(f"Loaded {len(rows)} rows from {csv_path}")
        if args.dry_run:
            print("Dry run — skipping database upsert")
            return 0
        count = upsert_rows(rows)
        print(f"Upserted {count} rows into seah_service_providers")
        return 0

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_file():
        raise SystemExit(f"Workbook not found: {xlsx_path}")

    rows = parse_workbook(xlsx_path)
    if not rows:
        raise SystemExit("No provider rows parsed from workbook")

    write_csv(rows, csv_path)
    print(f"Wrote {len(rows)} rows to {csv_path}")

    if args.dry_run:
        print("Dry run — skipping database upsert")
        return 0

    count = upsert_rows(rows)
    print(f"Upserted {count} rows into seah_service_providers")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
