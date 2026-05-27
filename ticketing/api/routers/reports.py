"""
Operational and quarterly GRM reports.
"""
import io
from datetime import date, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ticketing.api.dependencies import CurrentUser, get_current_user, get_db
from ticketing.api.schemas.reports import (
    QuarterlyAssignmentCreate,
    QuarterlyAssignmentOut,
    QuarterlyAssignmentUpdate,
    QuarterlyPlanResponse,
    QuarterlyReportLibraryCreate,
    QuarterlyReportLibraryItem,
    QuarterlyReportSchedule,
    QuarterlyReportTemplate,
    QuarterlyRolePlan,
    QuarterlyScheduleUpdate,
    ReportBuildRequest,
    ReportLimitsInfo,
    ReportQueryResponse,
    ReportSectionBlock,
    ReportSummaryResponse,
)
from ticketing.services.report_summary import build_report_summary
from ticketing.services.quarterly_library import (
    create_library_item,
    delete_library_item,
    get_library_item,
    list_library,
)
from ticketing.services.quarterly_assignments import (
    create_assignments_for_roles,
    delete_assignment,
    load_schedule,
    plan_summary,
    quarter_key_from_date,
    save_schedule,
    update_assignment,
)
from ticketing.services.report_limits import (
    check_export_rate_limit,
    load_report_limits,
    log_report_export,
)
from ticketing.services.pivot_table import (
    AGGREGATIONS,
    DIMENSION_FIELDS,
    MEASURE_FIELDS,
    build_pivot_table,
)
from ticketing.services.report_export import (
    summary_export_filename,
    summary_export_row_count,
    summary_workbook_bytes,
)
from ticketing.services.report_rows import (
    DEFAULT_REPORT_COLUMNS,
    FIELD_LABELS,
    GROUP_BY_KEYS,
    MAX_EXPORT_ROWS,
    aggregate_rows,
    build_xlsx_workbook,
    load_report_rows,
    project_row,
    split_sections,
)

router = APIRouter()


def _parse_id_list(value: Optional[str]) -> list[str] | None:
    if not value or not value.strip():
        return None
    return [x.strip() for x in value.split(",") if x.strip()]


def _field_catalog() -> list[dict[str, str]]:
    return [{"key": k, "label": FIELD_LABELS[k]} for k in FIELD_LABELS]


def _limits_info(db: Session) -> ReportLimitsInfo:
    lim = load_report_limits(db)
    allowed = lim.get("allowed_recipient_roles")
    return ReportLimitsInfo(
        max_export_rows=lim["max_export_rows"],
        max_exports_per_user_per_hour=lim["max_exports_per_user_per_hour"],
        max_reports_per_role_per_quarter=int(lim.get("max_reports_per_role_per_quarter", 3)),
        quarterly_email_enabled=bool(lim.get("quarterly_email_enabled", True)),
        allowed_recipient_roles=list(allowed) if allowed else None,
    )


def _guard_export(db: Session, current_user: CurrentUser, export_kind: str, row_count: int) -> None:
    try:
        check_export_rate_limit(db, current_user.user_id)
    except ValueError as exc:
        raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, detail=str(exc)) from exc
    lim = load_report_limits(db)
    cap = lim["max_export_rows"]
    if row_count > cap:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail=f"Export limited to {cap} rows. Narrow filters.",
        )
    log_report_export(db, current_user.user_id, export_kind=export_kind, row_count=row_count)


@router.get(
    "/reports/query",
    response_model=ReportQueryResponse,
    summary="Operational report — four sections with filters",
)
def query_report(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    project_ids: Optional[str] = Query(None, description="Comma-separated project_id values"),
    package_ids: Optional[str] = Query(None, description="Comma-separated package_id values"),
    location_codes: Optional[str] = Query(None, description="Comma-separated location_code values"),
    include_seah: bool = Query(False),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
) -> ReportQueryResponse:
    if date_from is None:
        today = date.today()
        q = (today.month - 1) // 3
        date_from = date(today.year, q * 3 + 1, 1)
    if date_to is None:
        date_to = date.today()

    if include_seah and not current_user.can_see_seah:
        include_seah = False

    rows = load_report_rows(
        db,
        current_user,
        date_from=date_from,
        date_to=date_to,
        project_ids=_parse_id_list(project_ids),
        package_ids=_parse_id_list(package_ids),
        location_codes=_parse_id_list(location_codes),
        include_seah=include_seah,
    )
    sections_raw = split_sections(rows)
    columns = DEFAULT_REPORT_COLUMNS

    def _page_slice(items: list[dict]) -> tuple[list[dict], int]:
        start = (page - 1) * page_size
        sliced = items[start : start + page_size]
        public = []
        for r in sliced:
            pr = project_row(r, columns)
            pr["ticket_id"] = r.get("ticket_id")
            public.append(pr)
        return public, len(items)

    sections_out: dict[str, ReportSectionBlock] = {}
    summary = {"total": len(rows), "resolved": 0, "high": 0, "overdue": 0, "other": 0}
    for key in ("resolved", "high", "overdue", "other"):
        items, total = _page_slice(sections_raw[key])
        sections_out[key] = ReportSectionBlock(items=items, total=total)
        summary[key] = total

    return ReportQueryResponse(
        filters={
            "date_from": str(date_from),
            "date_to": str(date_to),
            "project_ids": _parse_id_list(project_ids) or [],
            "package_ids": _parse_id_list(package_ids) or [],
            "location_codes": _parse_id_list(location_codes) or [],
            "include_seah": include_seah,
        },
        summary=summary,
        columns=columns,
        column_labels={c: FIELD_LABELS.get(c, c) for c in columns},
        sections=sections_out,
        field_catalog=_field_catalog(),
    )


def _summary_query_years(years: Optional[str]) -> list[int] | None:
    if not years:
        return None
    return [int(y.strip()) for y in years.split(",") if y.strip().isdigit()]


def _load_summary_payload(
    db: Session,
    current_user: CurrentUser,
    *,
    project_id: str,
    province_code: Optional[str],
    quarter_keys: Optional[str],
    years: Optional[str],
    chart_package_ids: Optional[str],
    include_seah: bool,
) -> dict:
    if include_seah and not current_user.can_see_seah:
        include_seah = False
    try:
        return build_report_summary(
            db,
            current_user,
            project_id=project_id,
            province_code=province_code,
            quarter_keys=_parse_id_list(quarter_keys),
            years=_summary_query_years(years),
            chart_package_ids=_parse_id_list(chart_package_ids),
            include_seah=include_seah,
        )
    except ValueError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


@router.get(
    "/reports/summary",
    response_model=ReportSummaryResponse,
    summary="Executive summary matrix + charts (ADB quarterly view)",
)
def report_summary(
    project_id: str = Query(..., description="Single project_id"),
    province_code: Optional[str] = Query(None),
    quarter_keys: Optional[str] = Query(None, description="Comma-separated e.g. 2026-Q1,2026-Q2"),
    years: Optional[str] = Query(None, description="Comma-separated years e.g. 2025,2026"),
    chart_package_ids: Optional[str] = Query(None, description="Optional package filter for charts"),
    include_seah: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
) -> ReportSummaryResponse:
    payload = _load_summary_payload(
        db,
        current_user,
        project_id=project_id,
        province_code=province_code,
        quarter_keys=quarter_keys,
        years=years,
        chart_package_ids=chart_package_ids,
        include_seah=include_seah,
    )
    return ReportSummaryResponse(**payload)


@router.get(
    "/reports/summary/export",
    summary="Download executive Summary report as XLSX (matrix + charts)",
    response_class=StreamingResponse,
)
def export_summary_report(
    project_id: str = Query(..., description="Single project_id"),
    province_code: Optional[str] = Query(None),
    quarter_keys: Optional[str] = Query(None, description="Comma-separated e.g. 2026-Q1,2026-Q2"),
    years: Optional[str] = Query(None, description="Comma-separated years e.g. 2025,2026"),
    chart_package_ids: Optional[str] = Query(None, description="Optional package filter for charts"),
    include_seah: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
) -> StreamingResponse:
    payload = _load_summary_payload(
        db,
        current_user,
        project_id=project_id,
        province_code=province_code,
        quarter_keys=quarter_keys,
        years=years,
        chart_package_ids=chart_package_ids,
        include_seah=include_seah,
    )
    row_count = summary_export_row_count(payload)
    _guard_export(db, current_user, "export_summary", row_count)
    xlsx_bytes = summary_workbook_bytes(payload)
    filename = summary_export_filename(payload)
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post(
    "/reports/build",
    summary="Report builder — custom columns, optional group-by",
)
def build_report(
    body: ReportBuildRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    include_seah = body.include_seah
    if include_seah and not current_user.can_see_seah:
        include_seah = False

    rows = load_report_rows(
        db,
        current_user,
        date_from=body.date_from,
        date_to=body.date_to,
        project_ids=body.project_ids or None,
        package_ids=body.package_ids or None,
        location_codes=body.location_codes or None,
        include_seah=include_seah,
    )
    public_rows = [
        {k: v for k, v in row.items() if not str(k).startswith("_")}
        for row in rows
    ]

    if body.pivot and (body.pivot.rows or body.pivot.columns or body.pivot.values):
        try:
            value_specs = [{"field": v.field, "agg": v.agg} for v in body.pivot.values]
            pivot_result = build_pivot_table(
                public_rows,
                row_dims=body.pivot.rows,
                col_dims=body.pivot.columns,
                value_specs=value_specs,
                filters=body.pivot.filters,
            )
        except ValueError as exc:
            raise HTTPException(400, detail=str(exc)) from exc

        out_cols = pivot_result["columns"]
        out_rows = pivot_result["rows"]
        if body.format == "xlsx":
            _guard_export(db, current_user, "build_pivot", len(public_rows))
            return _pivot_xlsx(pivot_result)
        start = (body.page - 1) * body.page_size
        sliced = out_rows[start : start + body.page_size]
        return {
            **pivot_result,
            "rows": sliced,
            "total": len(out_rows),
        }

    columns = body.columns or DEFAULT_REPORT_COLUMNS
    invalid = [c for c in columns if c not in FIELD_LABELS]
    if invalid:
        raise HTTPException(400, detail=f"Unknown columns: {invalid}")

    if body.group_by:
        if body.group_by not in GROUP_BY_KEYS:
            raise HTTPException(400, detail=f"group_by must be one of: {sorted(GROUP_BY_KEYS)}")
        agg = body.aggregate if body.aggregate != "none" else "count"
        grouped = aggregate_rows(public_rows, body.group_by, agg)
        if body.format == "xlsx":
            return _builder_xlsx(grouped, [body.group_by, "ticket_count"] + (
                ["avg_total_days"] if agg == "avg_total_days" else
                ["sum_total_days"] if agg == "sum_total_days" else []
            ))
        start = (body.page - 1) * body.page_size
        sliced = grouped[start : start + body.page_size]
        return {
            "columns": [body.group_by, "ticket_count"],
            "rows": sliced,
            "total": len(grouped),
            "grouped": True,
        }

    if body.format == "xlsx":
        _guard_export(db, current_user, "build_flat", len(public_rows))
        return _builder_xlsx(public_rows, columns)

    start = (body.page - 1) * body.page_size
    sliced = public_rows[start : start + body.page_size]
    return {
        "columns": columns,
        "column_labels": {c: FIELD_LABELS.get(c, c) for c in columns},
        "rows": sliced,
        "total": len(public_rows),
        "grouped": False,
    }


def _pivot_xlsx(pivot_result: dict) -> StreamingResponse:
    from ticketing.services.report_export import pivot_workbook_bytes

    buf = io.BytesIO(pivot_workbook_bytes(pivot_result))
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="grm_pivot_report.xlsx"'},
    )


def _builder_xlsx(rows: list[dict], columns: list[str]) -> StreamingResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    labels = [FIELD_LABELS.get(c, c) for c in columns]
    for col, label in enumerate(labels, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
    for r, row in enumerate(rows[:MAX_EXPORT_ROWS], 2):
        for c, key in enumerate(columns, 1):
            ws.cell(row=r, column=c, value=row.get(key, ""))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "grm_custom_report.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/reports/export",
    summary="Download GRM report as XLSX (four sheets)",
    response_class=StreamingResponse,
)
def export_report(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    organization_id: Optional[str] = Query(None, description="Legacy filter; prefer project/location filters"),
    project_ids: Optional[str] = Query(None),
    package_ids: Optional[str] = Query(None),
    location_codes: Optional[str] = Query(None),
    include_seah: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
) -> StreamingResponse:
    if date_from is None:
        date_from = date.today() - timedelta(days=90)
    if date_to is None:
        date_to = date.today()

    if include_seah and not current_user.can_see_seah:
        include_seah = False

    rows = load_report_rows(
        db,
        current_user,
        date_from=date_from,
        date_to=date_to,
        project_ids=_parse_id_list(project_ids),
        package_ids=_parse_id_list(package_ids),
        location_codes=_parse_id_list(location_codes),
        include_seah=include_seah,
    )
    if organization_id:
        rows = [r for r in rows if r.get("organization_id") == organization_id]

    sections = split_sections(rows)
    public_sections = {
        k: [project_row(r, DEFAULT_REPORT_COLUMNS) for r in v]
        for k, v in sections.items()
    }
    _guard_export(db, current_user, "export_overview", len(rows))
    xlsx_bytes = build_xlsx_workbook(public_sections, DEFAULT_REPORT_COLUMNS)

    filename = f"grm_report_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _assignment_out(raw: dict) -> QuarterlyAssignmentOut:
    tmpl = raw.get("template") or {}
    return QuarterlyAssignmentOut(
        id=raw["id"],
        quarter_key=raw["quarter_key"],
        role_key=raw["role_key"],
        name=raw.get("name") or "Quarterly report",
        template=QuarterlyReportTemplate(**tmpl),
        active=bool(raw.get("active", True)),
    )


def _plan_response(db: Session, quarter_key: str) -> QuarterlyPlanResponse:
    summary = plan_summary(db, quarter_key)
    roles = [
        QuarterlyRolePlan(
            role_key=r["role_key"],
            count=r["count"],
            max=r["max"],
            assignments=[_assignment_out(a) for a in r["assignments"]],
        )
        for r in summary["roles"]
    ]
    sched = summary["schedule"]
    return QuarterlyPlanResponse(
        quarter_key=summary["quarter_key"],
        max_per_role=summary["max_per_role"],
        schedule=QuarterlyReportSchedule(
            frequency="quarterly",
            day_of_month=int(sched.get("day_of_month", 5)),
        ),
        limits=_limits_info(db),
        roles=roles,
    )


@router.get(
    "/reports/quarterly-plan",
    response_model=QuarterlyPlanResponse,
    summary="Quarterly report plan by role (admin)",
)
def get_quarterly_plan(
    quarter_key: Optional[str] = Query(None, description="e.g. 2026-Q1; default current quarter"),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
) -> QuarterlyPlanResponse:
    if not current_user.is_admin:
        raise HTTPException(403, detail="Admin access required")
    qk = quarter_key or quarter_key_from_date()
    return _plan_response(db, qk)


@router.put(
    "/reports/quarterly-schedule",
    response_model=QuarterlyReportSchedule,
    summary="Update send day of month (admin)",
)
def put_quarterly_schedule(
    body: QuarterlyScheduleUpdate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
) -> QuarterlyReportSchedule:
    if not current_user.is_admin:
        raise HTTPException(403, detail="Admin access required")
    sched = save_schedule(
        db,
        {"frequency": "quarterly", "day_of_month": body.day_of_month},
        current_user.user_id,
    )
    return QuarterlyReportSchedule(
        frequency="quarterly",
        day_of_month=int(sched.get("day_of_month", 5)),
    )


@router.post(
    "/reports/quarterly-assignments",
    response_model=list[QuarterlyAssignmentOut],
    summary="Save report(s) for role(s) — one slot per role (admin)",
)
def post_quarterly_assignments(
    body: QuarterlyAssignmentCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
) -> list[QuarterlyAssignmentOut]:
    if not current_user.is_admin:
        raise HTTPException(403, detail="Admin access required")

    if body.library_id:
        item = get_library_item(db, body.library_id)
        if not item:
            raise HTTPException(404, detail="Saved report not found")
        name = item.get("name") or "Quarterly report"
        template = item.get("template") or {}
    elif body.template and body.name:
        name = body.name
        template = body.template.model_dump()
    else:
        raise HTTPException(
            400,
            detail="Provide library_id or both name and template",
        )

    try:
        created = create_assignments_for_roles(
            db,
            quarter_key=body.quarter_key,
            role_keys=body.role_keys,
            name=name,
            template=template,
            updated_by=current_user.user_id,
        )
    except ValueError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return [_assignment_out(a) for a in created]


def _library_out(raw: dict) -> QuarterlyReportLibraryItem:
    tmpl = raw.get("template") or {}
    return QuarterlyReportLibraryItem(
        id=raw["id"],
        name=raw.get("name") or "Untitled report",
        template=QuarterlyReportTemplate(**tmpl),
    )


@router.get(
    "/reports/quarterly-library",
    response_model=list[QuarterlyReportLibraryItem],
    summary="Saved report definitions for quarterly planning (admin)",
)
def get_quarterly_library(
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
) -> list[QuarterlyReportLibraryItem]:
    if not current_user.is_admin:
        raise HTTPException(403, detail="Admin access required")
    return [_library_out(x) for x in list_library(db)]


@router.post(
    "/reports/quarterly-library",
    response_model=QuarterlyReportLibraryItem,
    summary="Save a named report definition (admin)",
)
def post_quarterly_library(
    body: QuarterlyReportLibraryCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
) -> QuarterlyReportLibraryItem:
    if not current_user.is_admin:
        raise HTTPException(403, detail="Admin access required")
    item = create_library_item(
        db,
        name=body.name,
        template=body.template.model_dump(),
        updated_by=current_user.user_id,
    )
    return _library_out(item)


@router.delete(
    "/reports/quarterly-library/{item_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Delete a saved report definition (admin)",
)
def remove_quarterly_library_item(
    item_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
) -> None:
    if not current_user.is_admin:
        raise HTTPException(403, detail="Admin access required")
    try:
        delete_library_item(db, item_id, current_user.user_id)
    except ValueError as exc:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail=str(exc)) from exc


@router.patch(
    "/reports/quarterly-assignments/{assignment_id}",
    response_model=QuarterlyAssignmentOut,
    summary="Update a quarterly assignment (admin)",
)
def patch_quarterly_assignment(
    assignment_id: str,
    body: QuarterlyAssignmentUpdate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
) -> QuarterlyAssignmentOut:
    if not current_user.is_admin:
        raise HTTPException(403, detail="Admin access required")
    try:
        updated = update_assignment(
            db,
            assignment_id,
            name=body.name,
            template=body.template.model_dump() if body.template else None,
            active=body.active,
            updated_by=current_user.user_id,
        )
    except ValueError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return _assignment_out(updated)


@router.delete(
    "/reports/quarterly-assignments/{assignment_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Remove a quarterly assignment (admin)",
)
def remove_quarterly_assignment(
    assignment_id: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
) -> None:
    if not current_user.is_admin:
        raise HTTPException(403, detail="Admin access required")
    try:
        delete_assignment(db, assignment_id, current_user.user_id)
    except ValueError as exc:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail=str(exc)) from exc


@router.get(
    "/reports/fields",
    summary="Column catalog for report builder",
)
def report_fields(
    _current_user: CurrentUser = Depends(get_current_user),
) -> dict:
    return {
        "fields": _field_catalog(),
        "dimensions": [
            {"key": k, "label": FIELD_LABELS[k]}
            for k in sorted(DIMENSION_FIELDS)
            if k in FIELD_LABELS
        ],
        "measures": [
            {"key": k, "label": FIELD_LABELS[k]}
            for k in sorted(MEASURE_FIELDS)
            if k in FIELD_LABELS
        ],
        "group_by_options": sorted(GROUP_BY_KEYS),
        "aggregates": list(AGGREGATIONS),
        "default_columns": DEFAULT_REPORT_COLUMNS,
        "default_pivot": {
            "rows": ["project_name"],
            "columns": ["complaint_category"],
            "values": [{"field": "ticket_id", "agg": "count"}],
            "filters": {},
        },
    }
