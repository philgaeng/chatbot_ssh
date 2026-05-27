"""XLSX export helpers shared by HTTP routes and Celery quarterly reports."""
from __future__ import annotations

import io
import re
from typing import Any

from ticketing.services.report_rows import FIELD_LABELS, MAX_EXPORT_ROWS


def pivot_workbook_bytes(pivot_result: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pivot"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    group_fill = PatternFill("solid", fgColor="D6E4F0")
    group_font = Font(bold=True)

    columns = pivot_result["columns"]
    rows = pivot_result["rows"]
    header_rows = pivot_result.get("header_rows") or []
    data_start = 1
    row_dim_count = len(pivot_result.get("row_dims") or [])

    if header_rows:
        excel_row = 1
        for hi, hrow in enumerate(header_rows):
            excel_col = row_dim_count + 1 if hi > 0 and row_dim_count else 1
            for cell in hrow:
                c = ws.cell(row=excel_row, column=excel_col, value=cell.get("text", ""))
                rs = int(cell.get("row_span") or 1)
                cs = int(cell.get("col_span") or 1)
                if rs > 1 or cs > 1:
                    ws.merge_cells(
                        start_row=excel_row,
                        start_column=excel_col,
                        end_row=excel_row + rs - 1,
                        end_column=excel_col + cs - 1,
                    )
                if cell.get("kind") == "col_group":
                    c.fill = group_fill
                    c.font = group_font
                    c.alignment = Alignment(horizontal="center")
                elif cell.get("kind") == "row_dim":
                    c.fill = header_fill
                    c.font = header_font
                else:
                    c.font = Font(bold=True)
                excel_col += cs
            excel_row += 1
        data_start = excel_row
    else:
        labels = [FIELD_LABELS.get(c, c) for c in columns]
        for col, label in enumerate(labels, 1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.fill = header_fill
            cell.font = header_font
        data_start = 2

    for r, row in enumerate(rows[:MAX_EXPORT_ROWS], data_start):
        for c, key in enumerate(columns, 1):
            ws.cell(row=r, column=c, value=row.get(key, ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def summary_export_filename(summary: dict[str, Any]) -> str:
    filters = summary.get("filters") or {}
    project = re.sub(r"[^\w\-]+", "_", str(filters.get("project_name") or "project"))[:40]
    quarters = filters.get("quarter_keys") or []
    qpart = "-".join(quarters) if quarters else "summary"
    return f"grm-summary_{project}_{qpart}.xlsx"


def summary_workbook_bytes(summary: dict[str, Any]) -> bytes:
    """Single-sheet Summary XLSX: filter header rows, then the matrix table."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    group_fill = PatternFill("solid", fgColor="D6E4F0")
    group_font = Font(bold=True)
    title_fill = PatternFill("solid", fgColor="E8EEF4")

    ws = wb.active
    ws.title = "Summary"

    filters = summary.get("filters") or {}
    definitions = summary.get("definitions") or {}
    filter_rows = [
        ("Project", filters.get("project_name", "")),
        ("Project ID", filters.get("project_id", "")),
        ("Province", filters.get("province_code") or "All provinces"),
        ("Quarters", ", ".join(filters.get("quarter_keys") or [])),
        ("Period from", filters.get("period_union_from", "")),
        ("Period to", filters.get("period_union_to", "")),
        ("Include SEAH", "Yes" if filters.get("include_seah") else "No"),
        ("Chart packages", ", ".join(filters.get("chart_package_ids") or []) or "All"),
        ("Closed on time", definitions.get("closed_on_time", "")),
        ("Closed overdue", definitions.get("closed_overdue", "")),
    ]

    title_cell = ws.cell(row=1, column=1, value="GRM Executive Summary")
    title_cell.font = Font(bold=True, size=14)
    title_cell.fill = title_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    hdr_a = ws.cell(row=2, column=1, value="Filter")
    hdr_b = ws.cell(row=2, column=2, value="Value")
    hdr_a.font = Font(bold=True)
    hdr_b.font = Font(bold=True)

    for i, (label, val) in enumerate(filter_rows):
        r = 3 + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val)

    matrix = summary.get("matrix") or {}
    column_groups: list[dict[str, Any]] = matrix.get("column_groups") or []
    matrix_rows: list[dict[str, Any]] = matrix.get("rows") or []

    leaves: list[dict[str, str]] = []
    for group in column_groups:
        for child in group.get("children") or []:
            leaves.append(
                {
                    "key": child["key"],
                    "leaf": child.get("label", child["key"]),
                    "group": group.get("label", ""),
                }
            )

    matrix_header_row = 3 + len(filter_rows) + 1  # blank row after filters
    group_row = matrix_header_row + 1
    leaf_row = matrix_header_row + 2
    data_start_row = matrix_header_row + 3

    pkg_cell = ws.cell(row=group_row, column=1, value="Package")
    pkg_cell.fill = header_fill
    pkg_cell.font = header_font
    pkg_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=group_row, start_column=1, end_row=leaf_row, end_column=1)

    col_idx = 2
    for group in column_groups:
        children = group.get("children") or []
        span = len(children)
        if span == 0:
            continue
        cell = ws.cell(row=group_row, column=col_idx, value=group.get("label", ""))
        cell.fill = group_fill
        cell.font = group_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        if span > 1:
            ws.merge_cells(
                start_row=group_row,
                start_column=col_idx,
                end_row=group_row,
                end_column=col_idx + span - 1,
            )
        for i, child in enumerate(children):
            leaf_cell = ws.cell(row=leaf_row, column=col_idx + i, value=child.get("label", ""))
            leaf_cell.font = Font(bold=True)
            leaf_cell.alignment = Alignment(horizontal="center")
        col_idx += span

    for r_i, row in enumerate(matrix_rows, start=data_start_row):
        ws.cell(row=r_i, column=1, value=row.get("package_name", ""))
        cells = row.get("cells") or {}
        for c_i, leaf in enumerate(leaves, start=2):
            val = cells.get(leaf["key"], 0) or 0
            ws.cell(row=r_i, column=c_i, value=int(val))

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.freeze_panes = ws.cell(row=data_start_row, column=2).coordinate

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def summary_export_row_count(summary: dict[str, Any]) -> int:
    """Rough row count for export rate-limit logging."""
    matrix = summary.get("matrix") or {}
    return max(len(matrix.get("rows") or []), 1)
