"""
Operational report row builder — shared by /reports/query, /reports/build, and XLSX export.
Decisions: docs/ticketing_system/12_reports_and_report_builder.md §8 (product answers).
"""

from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from typing import Any, Optional
from zoneinfo import ZoneInfo

import sqlalchemy as sa
from sqlalchemy import and_, func, or_, select
from sqlalchemy.orm import Session

from ticketing.api.dependencies import CurrentUser
from ticketing.constants.resolution import resolution_category_label
from ticketing.engine.workflow_engine import compute_sla_deadline
from ticketing.models.officer_scope import OfficerScope
from ticketing.models.package import ProjectPackage
from ticketing.models.project import Project
from ticketing.models.ticket import Ticket, TicketEvent
from ticketing.models.ticket_viewer import TicketViewer
from ticketing.models.workflow import WorkflowStep
from ticketing.services.officer_jurisdiction import scope_ticket_filter

NEPAL_TZ = ZoneInfo("Asia/Kathmandu")
MAX_EXPORT_ROWS = 100
DEFAULT_PAGE_SIZE = 100

# Default columns for overview + quarterly export (§4)
DEFAULT_REPORT_COLUMNS = [
    "complaint_date",
    "grievance_id",
    "high_yn",
    "escalated_yn",
    "overdue_yn",
    "stage",
    "complaint_category",
    "days_in_stage",
    "total_days",
    "resolution_category",
    "status_code",
    "project_name",
    "package_label",
    "location_display",
]

FIELD_LABELS: dict[str, str] = {
    "ticket_id": "Complaints (count)",
    "complaint_date": "Date of complaint",
    "grievance_id": "Reference no.",
    "high_yn": "High (Y/N)",
    "escalated_yn": "Escalated (Y/N)",
    "overdue_yn": "Overdue (Y/N)",
    "stage": "Stage",
    "stage_level": "Level",
    "complaint_category": "Complaint category",
    "days_in_stage": "Days in stage",
    "total_days": "Total days",
    "resolution_category": "Resolution category",
    "status_code": "Status",
    "priority": "Priority",
    "project_name": "Project",
    "package_label": "Package",
    "location_display": "Location",
    "organization_id": "Organization",
    "is_seah": "Instance",
    "sla_breached": "SLA breached (Y/N)",
    "assigned_officer": "Assigned officer",
    "grievance_summary": "Summary",
    "report_bucket": "Section",
}

GROUP_BY_KEYS = frozenset({
    "project_name",
    "package_label",
    "location_display",
    "stage",
    "resolution_category",
    "complaint_category",
    "status_code",
    "is_seah",
})


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _to_nepal_date(dt: datetime | None) -> str:
    if dt is None:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(NEPAL_TZ).date().isoformat()


def _calendar_days_between(start: datetime | None, end: datetime | None) -> int | None:
    if start is None or end is None:
        return None
    if start.tzinfo is None:
        start = start.replace(tzinfo=timezone.utc)
    if end.tzinfo is None:
        end = end.replace(tzinfo=timezone.utc)
    d0 = start.astimezone(NEPAL_TZ).date()
    d1 = end.astimezone(NEPAL_TZ).date()
    return max(0, (d1 - d0).days)


def normalize_complaint_category(raw: str | None) -> str:
    """First category token, trimmed (§8.11 normalize)."""
    if not raw or not str(raw).strip():
        return ""
    text = str(raw).strip()
    for sep in (";", "|", "/"):
        text = text.replace(sep, ",")
    parts = [p.strip() for p in text.split(",") if p.strip()]
    return parts[0] if parts else ""


def _location_codes_with_descendants(db: Session, codes: list[str]) -> list[str]:
    if not codes:
        return []
    expanded: set[str] = set()
    for code in codes:
        rows = db.execute(
            sa.text("""
                WITH RECURSIVE subtree AS (
                    SELECT location_code FROM ticketing.locations WHERE location_code = :lc
                    UNION ALL
                    SELECT l.location_code
                    FROM ticketing.locations l
                    JOIN subtree s ON l.parent_location_code = s.location_code
                )
                SELECT location_code FROM subtree
            """),
            {"lc": code},
        ).scalars().all()
        expanded.update(rows)
    return list(expanded)


def _apply_officer_scope(q, db: Session, current_user: CurrentUser):
    if current_user.is_admin:
        return q
    scopes = db.execute(
        select(OfficerScope).where(OfficerScope.user_id == current_user.user_id)
    ).scalars().all()
    viewed_ticket_ids = select(TicketViewer.ticket_id).where(
        TicketViewer.user_id == current_user.user_id
    )
    if not scopes:
        return q.where(or_(
            Ticket.assigned_to_user_id == current_user.user_id,
            Ticket.ticket_id.in_(viewed_ticket_ids),
        ))
    scope_conditions = [scope_ticket_filter(db, scope) for scope in scopes]
    scope_conditions.append(Ticket.ticket_id.in_(viewed_ticket_ids))
    return q.where(or_(*scope_conditions))


def _period_filter(date_from: date, date_to: date):
    """Created in period OR still open (not resolved) as of reporting (§8.2 Q5)."""
    return or_(
        and_(
            func.date(Ticket.created_at) >= date_from,
            func.date(Ticket.created_at) <= date_to,
        ),
        and_(
            func.date(Ticket.created_at) < date_from,
            Ticket.status_code.notin_(("RESOLVED", "CLOSED")),
        ),
    )


def _resolved_in_period_filter(date_from: date, date_to: date, resolved_at_subq):
    return and_(
        Ticket.status_code.in_(("RESOLVED", "CLOSED")),
        func.date(resolved_at_subq) >= date_from,
        func.date(resolved_at_subq) <= date_to,
    )


def build_ticket_query(
    db: Session,
    current_user: CurrentUser,
    *,
    date_from: date,
    date_to: date,
    project_ids: list[str] | None = None,
    package_ids: list[str] | None = None,
    location_codes: list[str] | None = None,
    include_seah: bool = False,
) -> sa.sql.Select:
    q = select(Ticket).where(Ticket.is_deleted.is_(False), _period_filter(date_from, date_to))

    if not include_seah and not current_user.can_see_seah:
        q = q.where(Ticket.is_seah.is_(False))
    elif not include_seah:
        q = q.where(Ticket.is_seah.is_(False))
    # When include_seah=True and user can_see_seah, no extra filter

    q = _apply_officer_scope(q, db, current_user)

    if project_ids:
        q = q.where(or_(
            Ticket.project_id.in_(project_ids),
            Ticket.project_code.in_(
                select(Project.short_code).where(Project.project_id.in_(project_ids))
            ),
        ))
    if package_ids:
        q = q.where(Ticket.package_id.in_(package_ids))
    if location_codes:
        expanded = _location_codes_with_descendants(db, location_codes)
        if expanded:
            q = q.where(Ticket.location_code.in_(expanded))

    return q.order_by(Ticket.created_at.desc())


def _fetch_auxiliary_maps(
    db: Session,
    tickets: list[Ticket],
) -> tuple[
    dict[str, WorkflowStep],
    dict[str, str],
    dict[str, str],
    dict[str, datetime | None],
    set[str],
    dict[str, str | None],
]:
    step_ids = {t.current_step_id for t in tickets if t.current_step_id}
    step_map: dict[str, WorkflowStep] = {}
    if step_ids:
        for s in db.execute(select(WorkflowStep).where(WorkflowStep.step_id.in_(step_ids))).scalars():
            step_map[s.step_id] = s

    project_ids = {t.project_id for t in tickets if t.project_id}
    project_names: dict[str, str] = {}
    if project_ids:
        for p in db.execute(select(Project).where(Project.project_id.in_(project_ids))).scalars():
            project_names[p.project_id] = p.name

    package_ids = {t.package_id for t in tickets if t.package_id}
    package_labels: dict[str, str] = {}
    if package_ids:
        for pkg in db.execute(
            select(ProjectPackage).where(ProjectPackage.package_id.in_(package_ids))
        ).scalars():
            package_labels[pkg.package_id] = pkg.name

    ticket_ids = [t.ticket_id for t in tickets]
    resolved_at: dict[str, datetime | None] = {}
    resolution_cat: dict[str, str | None] = {}
    escalated_ids: set[str] = set()

    if ticket_ids:
        for row in db.execute(
            select(TicketEvent.ticket_id, TicketEvent.created_at, TicketEvent.payload)
            .where(
                TicketEvent.ticket_id.in_(ticket_ids),
                TicketEvent.event_type == "RESOLVED",
            )
            .order_by(TicketEvent.created_at.desc())
        ).all():
            tid = row[0]
            if tid not in resolved_at:
                resolved_at[tid] = row[1]
                payload = row[2] or {}
                resolution_cat[tid] = payload.get("resolution_category")

        escalated_ids = set(
            db.execute(
                select(TicketEvent.ticket_id)
                .where(
                    TicketEvent.ticket_id.in_(ticket_ids),
                    TicketEvent.event_type == "ESCALATED",
                )
                .distinct()
            ).scalars().all()
        )

    return step_map, project_names, package_labels, resolved_at, escalated_ids, resolution_cat


def _is_overdue_now(ticket: Ticket, step: WorkflowStep | None, now: datetime) -> bool:
    if ticket.current_overdue_episode_id:
        return True
    if ticket.sla_breached:
        return True
    deadline = compute_sla_deadline(ticket, step)
    if deadline is None:
        return False
    if deadline.tzinfo is None:
        deadline = deadline.replace(tzinfo=timezone.utc)
    return deadline < now


def _classify_sections(
    ticket: Ticket,
    *,
    step: WorkflowStep | None,
    resolved_at: datetime | None,
    date_from: date,
    date_to: date,
    now: datetime,
) -> list[str]:
    """Non-exclusive section tags (§8.5 Q13 overlap)."""
    sections: list[str] = []
    is_resolved = ticket.status_code in ("RESOLVED", "CLOSED")
    if is_resolved and ticket.created_at:
        created = ticket.created_at
        if created.tzinfo is None:
            created = created.replace(tzinfo=timezone.utc)
        cd = created.astimezone(NEPAL_TZ).date()
        if date_from <= cd <= date_to:
            sections.append("resolved")

    if not is_resolved:
        overdue = _is_overdue_now(ticket, step, now)
        is_high = (
            ticket.priority in ("HIGH", "CRITICAL")
            or ticket.is_seah
            or overdue
        )
        if is_high:
            sections.append("high")
        if overdue:
            sections.append("overdue")
        if not is_high and not overdue:
            sections.append("other")
    elif "resolved" not in sections and is_resolved:
        # Resolved outside period window — still in population via Q5 open rule? Unlikely if resolved.
        pass

    if not sections and not is_resolved:
        sections.append("other")

    return sections


def build_report_row(
    ticket: Ticket,
    *,
    step_map: dict[str, WorkflowStep],
    project_names: dict[str, str],
    package_labels: dict[str, str],
    resolved_at_map: dict[str, datetime | None],
    escalated_ids: set[str],
    resolution_cat_map: dict[str, str | None],
    date_from: date,
    date_to: date,
    now: datetime | None = None,
) -> dict[str, Any]:
    now = now or _now_utc()
    step = step_map.get(ticket.current_step_id) if ticket.current_step_id else None
    resolved_at = resolved_at_map.get(ticket.ticket_id)
    clock_end = resolved_at if ticket.status_code in ("RESOLVED", "CLOSED") and resolved_at else now
    stage_start = ticket.step_started_at or ticket.created_at

    overdue_now = _is_overdue_now(ticket, step, now)
    escalated = (
        ticket.ticket_id in escalated_ids
        or ticket.status_code == "ESCALATED"
    )
    is_high = (
        ticket.priority in ("HIGH", "CRITICAL")
        or ticket.is_seah
        or overdue_now
    )

    res_code = resolution_cat_map.get(ticket.ticket_id)
    res_label = resolution_category_label(res_code) if res_code else ""

    pkg_label = package_labels.get(ticket.package_id) if ticket.package_id else "(No package)"
    proj_name = ""
    if ticket.project_id:
        proj_name = project_names.get(ticket.project_id, "")
    elif ticket.project_code:
        proj_name = ticket.project_code

    sections = _classify_sections(
        ticket,
        step=step,
        resolved_at=resolved_at,
        date_from=date_from,
        date_to=date_to,
        now=now,
    )

    row: dict[str, Any] = {
        "ticket_id": ticket.ticket_id,
        "complaint_date": _to_nepal_date(ticket.created_at),
        "grievance_id": ticket.grievance_id,
        "high_yn": "Y" if is_high else "N",
        "escalated_yn": "Y" if escalated else "N",
        "overdue_yn": "Y" if overdue_now or ticket.sla_breached else "N",
        "stage": step.display_name if step else "",
        "stage_level": f"L{step.step_order}" if step else "",
        "complaint_category": normalize_complaint_category(ticket.grievance_categories),
        "days_in_stage": _calendar_days_between(stage_start, clock_end),
        "total_days": _calendar_days_between(ticket.created_at, clock_end),
        "resolution_category": res_label,
        "status_code": ticket.status_code,
        "priority": ticket.priority,
        "project_name": proj_name,
        "package_label": pkg_label,
        "location_display": ticket.grievance_location or ticket.location_code or "",
        "organization_id": ticket.organization_id,
        "is_seah": "SEAH" if ticket.is_seah else "Standard",
        "sla_breached": "Y" if ticket.sla_breached else "N",
        "assigned_officer": ticket.assigned_to_user_id or "",
        "grievance_summary": (ticket.grievance_summary or "")[:500],
        "report_bucket": sections[0] if len(sections) == 1 else ",".join(sections),
        "_sections": sections,
    }
    return row


def load_report_rows(
    db: Session,
    current_user: CurrentUser,
    *,
    date_from: date,
    date_to: date,
    project_ids: list[str] | None = None,
    package_ids: list[str] | None = None,
    location_codes: list[str] | None = None,
    include_seah: bool = False,
) -> list[dict[str, Any]]:
    q = build_ticket_query(
        db,
        current_user,
        date_from=date_from,
        date_to=date_to,
        project_ids=project_ids,
        package_ids=package_ids,
        location_codes=location_codes,
        include_seah=include_seah,
    )
    tickets = db.execute(q).scalars().all()
    if not tickets:
        return []
    aux = _fetch_auxiliary_maps(db, tickets)
    now = _now_utc()
    return [
        build_report_row(
            t,
            step_map=aux[0],
            project_names=aux[1],
            package_labels=aux[2],
            resolved_at_map=aux[3],
            escalated_ids=aux[4],
            resolution_cat_map=aux[5],
            date_from=date_from,
            date_to=date_to,
            now=now,
        )
        for t in tickets
    ]


def split_sections(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    out: dict[str, list[dict[str, Any]]] = {
        "resolved": [],
        "high": [],
        "overdue": [],
        "other": [],
    }
    for row in rows:
        for sec in row.get("_sections") or []:
            if sec in out:
                pub = {k: v for k, v in row.items() if not k.startswith("_")}
                out[sec].append(pub)
    return out


def project_row(row: dict[str, Any], columns: list[str]) -> dict[str, Any]:
    return {k: row.get(k, "") for k in columns if k in FIELD_LABELS or k in row}


def aggregate_rows(
    rows: list[dict[str, Any]],
    group_by: str,
    aggregate: str,
) -> list[dict[str, Any]]:
    """Group-by with count / avg_total_days (§8.10 basic functions)."""
    from collections import defaultdict

    buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        key = str(row.get(group_by) or "(blank)")
        buckets[key].append(row)

    result: list[dict[str, Any]] = []
    for key, items in sorted(buckets.items(), key=lambda x: x[0]):
        out: dict[str, Any] = {group_by: key}
        if aggregate == "count":
            out["ticket_count"] = len(items)
        elif aggregate == "avg_total_days":
            vals = [i.get("total_days") for i in items if isinstance(i.get("total_days"), int)]
            out["ticket_count"] = len(items)
            out["avg_total_days"] = round(sum(vals) / len(vals), 1) if vals else None
        elif aggregate == "sum_total_days":
            vals = [i.get("total_days") for i in items if isinstance(i.get("total_days"), int)]
            out["ticket_count"] = len(items)
            out["sum_total_days"] = sum(vals) if vals else 0
        else:
            out["ticket_count"] = len(items)
        result.append(out)
    return result


def build_xlsx_workbook(
    sections: dict[str, list[dict[str, Any]]],
    columns: list[str],
) -> bytes:
    import io

    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    labels = [FIELD_LABELS.get(c, c) for c in columns]

    for sheet_name, key in [
        ("Resolved", "resolved"),
        ("High", "high"),
        ("Overdue", "overdue"),
        ("Others", "other"),
    ]:
        ws = wb.create_sheet(title=sheet_name[:31])
        for col, label in enumerate(labels, 1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.fill = header_fill
            cell.font = header_font
        for r, row in enumerate(sections.get(key, [])[:MAX_EXPORT_ROWS], 2):
            for c, col_key in enumerate(columns, 1):
                ws.cell(row=r, column=c, value=row.get(col_key, ""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
